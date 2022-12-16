import os

from tqdm import tqdm

from openpyxl import load_workbook

from ukr_np_regex import regex_ukr, regex_np
from api_request import send_request_to_np, send_request_to_ukr


def read_excel(path_file):
    """ read data in excel """
    wb = load_workbook(filename=path_file)
    sheet = wb.active
    for i in tqdm(range(2, sheet.max_row + 1), position=0):
        col_f = sheet[f'H{i}'].value
        if col_f is not None:
            if regex_np(ttn=col_f) is not None:
                try:
                    response = send_request_to_np(ttn=f"{col_f}")
                    status = response.get('data')[0].get('Status')
                except Exception:
                    continue
            else:
                if regex_ukr(ttn=col_f) is not None:
                    try:
                        response = send_request_to_ukr(ttn=f"{col_f}")
                        status = response.get('eventName')
                    except Exception:
                        continue
                else:
                    col_f = f"0{col_f}"
                    try:
                        response = send_request_to_ukr(ttn=f"{col_f}")
                        status = response.get('eventName')

                    except Exception:
                        continue
        else:
            # print("empty field")
            status = ''
        sheet[f'K{i}'] = status

    name = os.path.split(path_file)[-1]

    wb.save(f"data/with_status/{name}")


def main():
    path_file = input("Enter path to file: ")
    read_excel(path_file=path_file)


if __name__ == "__main__":
    main()


